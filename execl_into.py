import os

import openpyxl

from utils.db import ConnectDatabase

# pat = r'C:\Users\86138\Desktop\20200905房小团\pdf_parse\utils\002pdf_excel.xlsx'

db = ConnectDatabase('ohop')


def readfile():
    rootdir = r'.\values'
    list = os.listdir(rootdir)  # 列出文件夹下所有的目录与文件
    for i in range(0, len(list)):
        # for i in range(3, 4):
        pos = os.path.join(rootdir, list[i])
        # print(pos)
        if os.path.isfile(pos):
            # 你想对文件的操作
            pt = os.path.abspath(pos)  # 绝对路径
            # print(pt)
            basename = os.path.basename(pos)
            name_id = basename.split('_')[0].split('.')[0]
            print(name_id)  # pid

            wb = openpyxl.load_workbook(pt)
            ws = wb.active  # 当前活跃的表单
            try:
                for row in ws[2:12]:  # 打印 2-5行中所有单元格中的值
                    try:
                        a = row[0].value
                        if a == '楼栋':
                            continue
                        else:
                            aa = a.split('幢')
                            # print(b)
                            if len(aa) == 1:
                                aaa = ['', aa[0]]
                                # dong = aaa[0]
                                # yuan = aaa[1].replace('单元', '')
                            else:
                                aaa = aa
                        dong = aaa[0]
                        yuan = aaa[1].replace('单元', '')

                    except:
                        dong = ''
                        yuan = ''
                    try:
                        b = row[1].value
                    except:
                        b = ''
                    try:
                        c = row[2].value
                    except:
                        c = ''
                    try:
                        d = row[3].value
                    except:
                        d = ''
                    try:
                        e = row[4].value
                    except:
                        e = ''
                    try:
                        f = row[5].value
                    except:
                        f = ''
                    try:
                        g = row[6].value
                    except:
                        g = ''
                    try:
                        h = row[7].value
                    except:
                        h = ''
                    sql = "insert into one_house_one_price(pid,house_dong,house_yuan,house_door,j_area," \
                          "tj_area,give_house,unit_price,all_price,sale_statue) values('{}','{}','{}'," \
                          "'{}','{}','{}','{}','{}','{}','{}');".format(name_id, dong, yuan, b, c, d, e, f, g, h)
                    db.run(sql)
            except:
                print(pt)
                # continue


if __name__ == '__main__':
    readfile()

    # db.close()
    # pt = r'C:\Users\86138\Desktop\20200905房小团\one_house_one_price\values\20201205221650251413.xlsx'
    # wb = openpyxl.load_workbook(pt)
    # ws = wb.active  # 当前活跃的表单
    # name_id = '20201205221650251413'
    # try:
    #     for row in ws[2:12]:  # 打印 2-5行中所有单元格中的值
    #         try:
    #             a = row[0].value
    #             if a == '楼栋':
    #                 continue
    #             else:
    #                 aa = a.split('幢')
    #                 # print(b)
    #                 if len(aa) == 1:
    #                     aaa = ['' , aa[0]]
    #                     dong = aaa[0]
    #                     yuan = aaa[1].replace('单元','')
    #                 else:
    #                     aaa = aa
    #             dong = aaa[0]
    #             yuan = aaa[1].replace('单元','')
    #
    #         except:
    #             dong = ''
    #             yuan = ''
    #         try:
    #             b = row[1].value
    #         except:
    #             b = ''
    #         try:
    #             c = row[2].value
    #         except:
    #             c = ''
    #         try:
    #             d = row[3].value
    #         except:
    #             d = ''
    #         try:
    #             e = row[4].value
    #         except:
    #             e = ''
    #         try:
    #             f = row[5].value
    #         except:
    #             f = ''
    #         try:
    #             g = row[6].value
    #         except:
    #             g = ''
    #         try:
    #             h = row[7].value
    #         except:
    #             h = ''
    #         sql = "insert into one_house_one_price(pid,house_dong,house_yuan,house_door,j_area," \
    #               "tj_area,give_house,unit_price,all_price,sale_statue) values('{}','{}','{}'," \
    #               "'{}','{}','{}','{}','{}','{}','{}');".format(name_id, dong, yuan,b, c,d, e, f, g,h)
    #         db.run(sql)
    # except:
    #         print(pt)
    #         # continue
